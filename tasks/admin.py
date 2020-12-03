from django.contrib.gis import admin
from django.http import HttpResponse
from django.utils import timezone
from import_export import resources, fields
from import_export.admin import ImportExportModelAdmin, ImportExportActionModelAdmin
from import_export.widgets import ForeignKeyWidget, ManyToManyWidget
from .models import *
from utils.models import *
from django.contrib.gis.db import models
from mapwidgets.widgets import GooglePointFieldInlineWidget
from leaflet.admin import LeafletGeoAdminMixin
from advanced_filters.admin import AdminAdvancedFiltersMixin

class OriginItemInline(LeafletGeoAdminMixin, admin.StackedInline):
    model = Origin
    extra = 1

class DestinationItemInline(LeafletGeoAdminMixin, admin.StackedInline):
    model = Destination
    extra = 1

class PackageItemInline(admin.TabularInline):
    model = OrderPackage
    extra = 1

class OrderStatusInline(admin.TabularInline):
    model = OrderStatus
    #readonly_fields = ['st_update']
    extra = 1
    can_delete = False
    can_add = False
    verbose_name = 'Order Status Records'
    verbose_name_plural = 'Order Status Records'

    def get_readonly_fields(self, request, obj=None):
        return [f.name for f in self.model._meta.fields]

class OrderAdmin(admin.ModelAdmin):

    model = Order
    icon_name = 'airport_shuttle'
    fieldsets = (('TITULO', {
                    'fields': ('id', 'title', 'description', 'ord_price')
                    }),
                 ('CLIENTE', {
                     'fields': (('client', 'request_id'),)
                 }),
                 ('STATUS', {
                     'fields': ('last_status', 'last_provider', 'last_driver', 'last_location', 'last_description')
                 }),
                 ('VENTANA DE TIEMPO', {
                     'fields': ('created_at', 'start_time', 'end_time', 'duration')
                 }),
                 )
                 
    readonly_fields = ['id', 'created_at']
    print([f.name for f in model._meta.fields if f not in ['last_status', 'last_provider', 'last_driver', 'last_location', 'last_description']])
    massadmin_exclude = [f.name for f in model._meta.fields if f.name not in ('last_status', 'last_provider', 'last_driver', 'last_location', 'last_description')]
    # list of fields to display in django admin
    list_display = ('title', 'client', 'created_at', 'start_time', 'end_time', 'status')
    #list_display_links = ('title', 'last_status')

    # if you want django admin to show the search bar, just add this line
    search_fields = ("title__startswith", )
    filter_horizontal = ()
    list_filter = ('client', 'start_time', 'end_time', 'last_status')

    # inlines
    inlines = (OriginItemInline, DestinationItemInline, PackageItemInline, OrderStatusInline,)

    def export_xlsx(modeladmin, request, queryset):

        import openpyxl
        import re
        from openpyxl.utils import get_column_letter
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = re.sub(' ', '_', '{}-export.xlsx'.format(timezone.localtime()))
        response['Content-Disposition'] = 'attachment; filename={}'.format(filename)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MyModel"
        row_num = 0
        columns = [
            (u"Task Title", 30),
            (u"Start Date", 20),
            (u"End Date", 20),
            (u"Address", 80),
            (u"Latitude", 20),
            (u"Longitude", 20),
            (u"Description", 30),
            (u"Order ID", 32),
            (u"Performer", 32),
            (u"Price", 20),
            (u"Client", 20),
            (u"Req_id", 32),
        ]
        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            # c.style.font.bold = True
            # set column width
            ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]
        for obj in queryset:
            dest = Destination.objects.filter(order=obj.id).values()[0]
            status = OrderStatus.objects.filter(order=obj.id).last()
            wk = ('street', 'house_num', 'suburb', 'city', 'province', 'country', 'pos_code')
            address = re.sub(',', '', ', '.join(str(value) for value in dict(zip(wk, [dest[k] for k in wk])).values() if value), 1)
            row_num += 1
            row = [
                dest['name'],
                obj.start_time,
                obj.end_time,
                address,
                dest['location'].y,
                dest['location'].x,
                obj.description,
                str(obj.id),
                status.provider.prov_name,
                obj.ord_price,
                obj.client.client_name,
                obj.request_id,
            ]
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                # c.style.alignment.wrap_text = True
        wb.save(response)
        return response

    actions = [export_xlsx]


admin.site.register(Order, OrderAdmin)

admin.site.site_header = 'PackGO Admin'
admin.site.site_title = 'PackGO Logistica'